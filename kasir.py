import streamlit as st
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import pandas as pd
import io

# Fungsi untuk membaca data barang dan saldo dari file Excel
def baca_dari_excel(nama_file):
    try:
        wb = openpyxl.load_workbook(nama_file)
        ws_barang = wb["Data Barang"]
        ws_saldo = wb["Saldo"]
        
        # Membaca data barang dari setiap baris
        barang = {}
        for row in range(2, ws_barang.max_row + 1):
            id_barang = ws_barang.cell(row=row, column=1).value
            nama_barang = ws_barang.cell(row=row, column=2).value
            harga_barang = ws_barang.cell(row=row, column=3).value
            stok_barang = ws_barang.cell(row=row, column=4).value
            
            barang[id_barang] = {
                'nama': nama_barang,
                'harga': harga_barang,
                'stok': stok_barang
            }

        # Membaca saldo
        saldo = ws_saldo.cell(row=1, column=1).value
        
        return barang, saldo
        
    except FileNotFoundError:
        st.error(f"File '{nama_file}' tidak ditemukan. Membuat dictionary barang kosong.")
        return {}, 0

# Fungsi untuk menyimpan data barang dan saldo ke dalam file Excel
def simpan_ke_excel(barang, saldo, nama_file):
    wb = Workbook()
    
    # Menyimpan data barang
    ws_barang = wb.active
    ws_barang.title = "Data Barang"
    ws_barang['A1'] = "ID"
    ws_barang['B1'] = "Nama Barang"
    ws_barang['C1'] = "Harga"
    ws_barang['D1'] = "Stok"
    
    row = 2
    for id_barang, detail_barang in barang.items():
        ws_barang.cell(row=row, column=1).value = id_barang
        ws_barang.cell(row=row, column=2).value = detail_barang['nama']
        ws_barang.cell(row=row, column=3).value = detail_barang['harga']
        ws_barang.cell(row=row, column=4).value = detail_barang['stok']
        row += 1

    # Menyimpan saldo
    ws_saldo = wb.create_sheet("Saldo")
    ws_saldo.cell(row=1, column=1).value = saldo
    
    wb.save(nama_file)
    st.success(f"Data barang dan saldo berhasil disimpan ke {nama_file}")

# Fungsi untuk format angka dengan pemisah ribuan
def format_angka(angka):
    return "{:,}".format(angka).replace(",", ".")

# Fungsi untuk menampilkan semua barang di Streamlit
def tampilkan_semua_barang(barang):
    st.subheader("Daftar Barang")
    
    # Membuat data frame dari dictionary barang
    data_barang = {
        "ID": [],
        "Nama Barang": [],
        "Harga": [],
        "Stok": []
    }
    
    for id_barang, detail_barang in barang.items():
        data_barang["ID"].append(id_barang)
        data_barang["Nama Barang"].append(detail_barang['nama'])
        data_barang["Harga"].append(format_angka(detail_barang['harga']))  # Format harga
        data_barang["Stok"].append(detail_barang['stok'])
    
    # Menampilkan data barang dalam format tabel
    df_barang = pd.DataFrame(data_barang)
    st.dataframe(df_barang, use_container_width=True)

# Fungsi untuk menambah barang baru
def tambah_barang(barang, saldo):
    st.subheader("Tambah Barang Baru")
    nama = st.text_input("Masukkan nama barang:")
    harga = st.number_input("Masukkan harga barang:", min_value=0)
    stok = st.number_input("Masukkan stok barang:", min_value=0, step=1)
    
    if st.button("Tambah Barang"):
        total_harga = harga * stok
        if total_harga > saldo:
            st.error("Saldo tidak mencukupi untuk menambahkan barang ini.")
        else:
            id_barang_baru = max(barang.keys(), default=0) + 1
            barang[id_barang_baru] = {'nama': nama, 'harga': harga, 'stok': stok}
            saldo -= total_harga  # Kurangi saldo dengan total harga penambahan barang
            simpan_ke_excel(barang, saldo, "data_barang.xlsx")  # Simpan data barang yang sudah diupdate ke file Excel
            st.success(f"Barang '{nama}' dengan ID {id_barang_baru} berhasil ditambahkan.")
            st.session_state.barang = barang  # Update session state
            st.session_state.saldo = saldo  # Update session state

# Fungsi untuk mencari barang berdasarkan ID
def cari_barang(barang):
    st.subheader("Cari Barang")
    id_barang = st.number_input("Masukkan ID barang yang ingin dicari:", min_value=1, step=1)
    
    if st.button("Cari Barang"):
        if id_barang in barang:
            detail = barang[id_barang]
            st.write(f"ID: {id_barang}")
            st.write(f"Nama Barang: {detail['nama']}")
            st.write(f"Harga: Rp {format_angka(detail['harga'])}")  # Format harga
            st.write(f"Stok: {detail['stok']}")
        else:
            st.error(f"Barang dengan ID {id_barang} tidak ditemukan.")

# Fungsi untuk memodifikasi atau menghapus barang
def modifikasi_barang(barang, saldo):
    st.subheader("Modifikasi Barang")
    
    id_barang = st.number_input("Masukkan ID barang yang ingin dimodifikasi atau dihapus:", min_value=1, step=1)
    
    if id_barang in barang:
        # Menampilkan detail barang yang dipilih
        st.write(f"ID: {id_barang}")
        st.write(f"Nama Barang: {barang[id_barang]['nama']}")
        st.write(f"Harga: Rp {format_angka(barang[id_barang]['harga'])}")
        st.write(f"Stok: {barang[id_barang]['stok']}")
        
        nama_baru = st.text_input("Masukkan nama baru:", value=barang[id_barang]['nama'])
        harga_baru = st.number_input("Masukkan harga baru:", min_value=0, value=barang[id_barang]['harga'])
        stok_baru = st.number_input("Masukkan stok baru:", min_value=0, step=1, value=barang[id_barang]['stok'])
        
        # Hitung perubahan stok dan harga
        perubahan_stok = stok_baru - barang[id_barang]['stok']
        total_harga = perubahan_stok * harga_baru
        
        # Tampilkan informasi tentang perubahan
        st.write(f"Perubahan stok: {perubahan_stok}")
        st.write(f"Total biaya untuk modifikasi stok: Rp {format_angka(total_harga)}")
        
        if st.button("Simpan Perubahan"):
            if total_harga > saldo:
                st.error("Saldo Anda tidak mencukupi untuk modifikasi stok barang.")
            else:
                barang[id_barang]['nama'] = nama_baru
                barang[id_barang]['harga'] = harga_baru
                saldo -= total_harga  # Kurangi saldo dengan total harga modifikasi stok
                barang[id_barang]['stok'] = stok_baru
                simpan_ke_excel(barang, saldo, "data_barang.xlsx")  # Simpan data barang yang sudah diupdate ke file Excel
                st.success(f"Barang dengan ID {id_barang} berhasil dimodifikasi.")
                st.session_state.barang = barang  # Update session state
                st.session_state.saldo = saldo  # Update session state

        if st.button("Hapus Barang"):
            # Hapus barang
            del barang[id_barang]
            
            # Urutkan kembali ID barang
            barang_baru = {}
            next_id = 1
            for k, v in sorted(barang.items()):
                barang_baru[next_id] = v
                next_id += 1
            
            barang = barang_baru  # Update dictionary barang dengan ID yang sudah diatur ulang
            
            simpan_ke_excel(barang, saldo, "data_barang.xlsx")  # Simpan data barang yang sudah diupdate ke file Excel
            st.success(f"Barang dengan ID {id_barang} berhasil dihapus.")
            st.session_state.barang = barang  # Update session state
            st.session_state.saldo = saldo  # Update session state
    else:
        st.error(f"Barang dengan ID {id_barang} tidak ditemukan.")

# Fungsi untuk melakukan pembelian barang
def beli_barang(barang, saldo):
    st.subheader("Beli Barang")
    total_belanja = 0
    transaksi = []  # untuk menyimpan detail transaksi
    
    # List untuk menyimpan ID barang yang ingin dibeli
    list_id_barang = []
    # Dict untuk menyimpan jumlah barang yang ingin dibeli dengan key ID barang
    jumlah_barang_dibeli = {}
    
    barang_list = [f"{barang[id_barang]['nama']} (ID: {id_barang})" for id_barang in barang]
    
    selected_barang = st.multiselect("Pilih barang yang ingin dibeli:", barang_list)
    
    for item in selected_barang:
        id_barang = int(item.split("(ID: ")[1][:-1])
        jumlah = st.number_input(f"Masukkan jumlah {barang[id_barang]['nama']} yang ingin dibeli:", min_value=1, step=1, key=f"jumlah_{id_barang}")
        if id_barang not in list_id_barang:
            list_id_barang.append(id_barang)
            jumlah_barang_dibeli[id_barang] = 0  # Set jumlah awal untuk barang yang baru diinput
        
        jumlah_barang_dibeli[id_barang] = jumlah  # Update jumlah barang yang ingin dibeli
    
    # Proses pembelian
    if list_id_barang:
        st.subheader("Daftar Barang yang Akan Dibeli")
        for id_barang in list_id_barang:
            if jumlah_barang_dibeli[id_barang] > 0:
                if jumlah_barang_dibeli[id_barang] > barang[id_barang]['stok']:
                    st.error(f"Stok tidak mencukupi untuk {barang[id_barang]['nama']}.")
                    return
                total_harga = barang[id_barang]['harga'] * jumlah_barang_dibeli[id_barang]
                st.write(f"Nama: {barang[id_barang]['nama']} (ID: {id_barang})")
                st.write(f"Stok Tersedia: {barang[id_barang]['stok']}")
                st.write(f"Harga Satuan: Rp {format_angka(barang[id_barang]['harga'])}")
                st.write(f"Jumlah: {jumlah_barang_dibeli[id_barang]}")
                st.write(f"Total: Rp {format_angka(total_harga)}")
                total_belanja += total_harga

        st.write(f"\nTotal Belanja: Rp {format_angka(total_belanja)}")
        st.write(f"Saldo Awal: Rp {format_angka(saldo)}")
        st.write(f"Saldo Setelah Pembelian: Rp {format_angka(saldo + total_belanja)}")

        if st.button("Konfirmasi Pembelian"):
            saldo += total_belanja  # Tambah saldo dengan total belanja
                
            # Buat struk pembelian
            struk = io.StringIO()
            struk.write("Struk Pembelian\n")
            struk.write("=================================\n")
            for id_barang in list_id_barang:
                if jumlah_barang_dibeli[id_barang] > 0:
                    total_harga = barang[id_barang]['harga'] * jumlah_barang_dibeli[id_barang]
                    struk.write(f"ID: {id_barang} - Nama: {barang[id_barang]['nama']} - Harga Satuan: Rp {format_angka(barang[id_barang]['harga'])} - Jumlah: {jumlah_barang_dibeli[id_barang]} - Total: Rp {format_angka(total_harga)}\n")
            struk.write("\n")
            struk.write(f"Saldo Awal: Rp {format_angka(saldo - total_belanja)}\n")
            struk.write(f"Saldo Setelah Pembelian: Rp {format_angka(saldo)}\n")
            
            st.download_button("Unduh Struk Pembelian", data=struk.getvalue(), file_name="struk_pembelian.txt")
            
            # Update stok barang
            for id_barang in list_id_barang:
                barang[id_barang]['stok'] -= jumlah_barang_dibeli[id_barang]
            
            simpan_ke_excel(barang, saldo, "data_barang.xlsx")  # Simpan data barang yang sudah diupdate ke file Excel
            st.success(f"Pembelian berhasil dilakukan. Saldo baru: Rp {format_angka(saldo)}.")
            st.session_state.barang = barang  # Update session state
            st.session_state.saldo = saldo  # Update session state
    else:
        st.warning("Silakan pilih barang yang ingin dibeli.")

# Fungsi untuk mengedit saldo
def edit_saldo(saldo):
    st.subheader("Edit Saldo")
    tambah_saldo = st.number_input("Masukkan jumlah saldo yang ingin ditambahkan:", min_value=0, step=1000)
    if st.button("Tambah Saldo"):
        saldo += tambah_saldo
        simpan_ke_excel(st.session_state.barang, saldo, "data_barang.xlsx")  # Simpan data barang dan saldo yang sudah diupdate ke file Excel
        st.success(f"Saldo berhasil ditambahkan. Saldo baru: Rp {format_angka(saldo)}.")
        st.session_state.saldo = saldo  # Update session state

# Fungsi utama untuk menu aplikasi
def menu_utama():
    st.title("Aplikasi Kasir")
    
    # Membaca data dari file Excel
    barang, saldo = baca_dari_excel("data_barang.xlsx")
    
    # Menyimpan barang dan saldo ke session state
    if 'barang' not in st.session_state:
        st.session_state.barang = barang
    if 'saldo' not in st.session_state:
        st.session_state.saldo = saldo
    
    # Menampilkan saldo saat ini
    st.write(f"Sisa Saldo: Rp {format_angka(st.session_state.saldo)}")

    # Menggunakan tombol untuk memilih menu
    menu = st.selectbox("Pilih Menu:", ["Tampilkan Barang", "Tambah Barang", "Cari Barang", "Modifikasi Barang", "Beli Barang", "Edit Saldo"])
    
    if menu == "Tampilkan Barang":
        tampilkan_semua_barang(st.session_state.barang)
        
    elif menu == "Tambah Barang":
        tambah_barang(st.session_state.barang, st.session_state.saldo)
        
    elif menu == "Cari Barang":
        cari_barang(st.session_state.barang)
        
    elif menu == "Modifikasi Barang":
        modifikasi_barang(st.session_state.barang, st.session_state.saldo)
        
    elif menu == "Beli Barang":
        beli_barang(st.session_state.barang, st.session_state.saldo)
        
    elif menu == "Edit Saldo":
        edit_saldo(st.session_state.saldo)

# Menjalankan fungsi utama
if __name__ == "__main__":
    menu_utama()